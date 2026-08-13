import os
import json
import csv
import io
import pypdf
import docx
import openpyxl

def extract_text_from_file(file_name: str, file_bytes: bytes) -> str:
    """
    Extracts plain text content from various file buffers based on file extensions.
    Supports PDF, DOCX, TXT, CSV, XLSX, MD, JSON.
    """
    ext = os.path.splitext(file_name.lower())[1]
    
    if ext == '.txt' or ext == '.md':
        return file_bytes.decode('utf-8', errors='ignore')
        
    elif ext == '.json':
        try:
            data = json.loads(file_bytes.decode('utf-8', errors='ignore'))
            return json.dumps(data, indent=2)
        except Exception as e:
            return f"Error parsing JSON: {str(e)}"
            
    elif ext == '.csv':
        try:
            stream = io.StringIO(file_bytes.decode('utf-8', errors='ignore'))
            reader = csv.reader(stream)
            lines = []
            for row in reader:
                lines.append(", ".join(row))
            return "\n".join(lines)
        except Exception as e:
            return f"Error parsing CSV: {str(e)}"
            
    elif ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    row_strs = [str(val) if val is not None else "" for val in row]
                    # Filter out empty rows
                    if any(row_strs):
                        lines.append(" | ".join(row_strs))
            return "\n".join(lines)
        except Exception as e:
            return f"Error parsing Excel: {str(e)}"
            
    elif ext == '.docx':
        try:
            doc = docx.Document(io.BytesIO(file_bytes))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)
        except Exception as e:
            return f"Error parsing Word DOCX: {str(e)}"
            
    elif ext == '.pdf':
        try:
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n".join(pages)
        except Exception as e:
            return f"Error parsing PDF: {str(e)}"
            
    else:
        raise ValueError(f"Unsupported file format: {ext}")

def chunk_text(text: str, chunk_size: int = 1200, overlap: int = 200) -> list[str]:
    """
    Splits text into chunks of specified sizes with overlaps to handle large documents.
    """
    if len(text) <= chunk_size:
        return [text]
        
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks

def retrieve_relevant_chunks(query: str, chunks: list[str], top_n: int = 3) -> str:
    """
    Performs a lightweight keyword frequency matching to select the top_n most 
    relevant chunks for the given question, avoiding context window bloats.
    """
    # Simple stop words to filter out of the query match
    stop_words = {
        'the', 'is', 'a', 'of', 'and', 'to', 'in', 'it', 'is', 'was', 'for', 'on', 'with', 
        'as', 'at', 'by', 'an', 'be', 'this', 'that', 'from', 'or', 'are', 'what', 'how', 
        'where', 'when', 'who', 'which', 'about', 'can', 'your', 'my', 'he', 'she', 'they'
    }
    
    # Tokenize query words
    query_words = [
        w.strip("?,.:;!\"'()[]").lower() 
        for w in query.split() 
        if w.strip("?,.:;!\"'()[]").lower() not in stop_words and len(w) > 2
    ]
    
    if not query_words:
        # If query has only common words, just return the first few chunks
        return "\n\n[=== NEXT CHUNK ===]\n\n".join(chunks[:top_n])
        
    chunk_scores = []
    for chunk in chunks:
        chunk_lower = chunk.lower()
        score = 0
        for word in query_words:
            # Count word matches inside the chunk text
            score += chunk_lower.count(word)
        chunk_scores.append((score, chunk))
        
    # Sort chunks descending by match score
    chunk_scores.sort(key=lambda x: x[0], reverse=True)
    
    # Return the selected top_n chunks joined together
    best_chunks = [item[1] for item in chunk_scores[:top_n]]
    return "\n\n[=== NEXT CHUNK ===]\n\n".join(best_chunks)
